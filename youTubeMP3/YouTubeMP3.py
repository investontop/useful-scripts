# import
import configparser
import os
from pytube import YouTube
from moviepy.editor import VideoFileClip
import openpyxl
from datetime import datetime


def initiate_env_var(whattoreturn):
    if whattoreturn == 'ConfigFile':
        return '01Config.ini'

# read configs
config = configparser.ConfigParser()        # instance
configFile = initiate_env_var('ConfigFile')
config.read(configFile)

def download_youtube_video(url, output_path, SongName):
    # Download YouTube video
    yt = YouTube(url)
    stream = yt.streams.filter(only_audio=True).first()
    stream.download(output_path=output_path, filename=SongName)

# def convert_to_mp3(input_path, output_path):
#     # Convert video to MP3
#     video_clip = VideoFileClip(input_path)
#     audio_clip = video_clip.audio
#     audio_clip.write_audiofile(output_path)
#     audio_clip.close()
#     video_clip.close()
#
# def download_youtube_video_as_mp3(url, output_path):
#     # Download and convert YouTube video to MP3
#     download_youtube_video(url, output_path)
#     input_path = output_path + "/temp.mp4"
#     output_mp3_path = output_path + "/audio.mp3"
#     convert_to_mp3(input_path, output_mp3_path)

def read_excel(file_path, pRow, pCol):
    data = sheet.cell(row=pRow, column=pCol).value
    return data

intend = '      '
current_filename = os.path.basename(__file__)
print('['+datetime.now().strftime("%Y-%m-%d %H:%M:%S")+'] [' + current_filename  + "] Started - Considering config: ["
      + configFile + "]")
print('Note: Exit Code 0 means Success')

# Variable assigning
sourcePath = config['YouTubeMP3']['sourcepath']
downloadto = config['YouTubeMP3']['downloadpath']
workBook = config['YouTubeMP3']['workBook']

workbookPath = os.path.join(sourcePath, workBook)

# print(workbookPath)
workbook = openpyxl.load_workbook(workbookPath)
sheet = workbook.active  # Assuming you want to read from the active sheet

# Get the total number of rows
# print(sheet.max_row)
# print(read_excel(workbookPath, 2, 2))

for row in sheet.iter_rows(min_row = 2, values_only=True):
    try:
        cell_value_b = row[1]  # Value from column B (index 1) #URL
        cell_value_c = row[2]  # Value from column C (index 2) #MovieName-Song
        language, movieName, SongName = cell_value_c.split("-")

        # Create the necessary language folder
        LanguageFolder = os.path.join(downloadto, language)
        if not os.path.exists(LanguageFolder):
            os.makedirs(LanguageFolder)

        # Create the necessary MovieFolder
        MovieFolder = os.path.join(LanguageFolder, movieName)
        if not os.path.exists(MovieFolder):
            os.makedirs(MovieFolder)

        if os.path.exists(os.path.join(MovieFolder, SongName+".mp3")):
            print(intend + SongName+".mp3 - exists")
        else:
            # download_youtube_video_as_mp3(cell_value_b, MovieFolder)
            download_youtube_video(cell_value_b, MovieFolder, SongName+".mp3")
            print(intend + SongName+".mp3 - downloaded")

    except Exception as e:
        print(intend + f"<<<error occurred>>>: {e}:  Row: {row}")

#download_youtube_video_as_mp3(cell_value_b, MovieFolder, SongName+.mp3)

print('['+datetime.now().strftime("%Y-%m-%d %H:%M:%S")+'] [' + current_filename  + "] Completed")